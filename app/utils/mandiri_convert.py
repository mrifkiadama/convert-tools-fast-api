import re
import camelot
import pdfplumber
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def classify_mandiri_transaction(row):
    nominal_val = str(row.get("Nominal", "")).strip()
    # Remove thousand separators (.) and replace decimal comma with dot
    clean_val = nominal_val.replace(".", "").replace(",", ".").replace(" ", "")

    try:
        # Try to parse float, ignore signs for now
        amount = float(clean_val.replace("+", "").replace("-", ""))
    except Exception:
        amount = 0.0

    # Check original sign presence for credit/debit classification
    if "+" in nominal_val:
        return pd.Series([amount, ""])
    elif "-" in nominal_val:
        return pd.Series(["", amount])
    else:
        # If no explicit sign, guess by amount sign or 0
        if amount > 0:
            return pd.Series([amount, ""])
        else:
            return pd.Series(["", ""])


async def extract_mandiri_transactions(pdf_path: str, export_type: str) -> BytesIO:
    output = BytesIO()

    header_lines = []
    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text()

    # Extract fields using regex (you can fine-tune this based on actual layout)
    name = re.search(r"Nama/Name\s*:\s*(.+?)\s+Periode", text)
    periode = re.search(r"Periode/Period\s*:\s*(\d{2}\s\w+\s\d{4})\s*-\s*(\d{2}\s\w+\s\d{4})", text)
    branch = re.search(r"Cabang/Branch\s*:\s*(.+?)\s+Dicetak", text)
    printed = re.search(r"Dicetak pada/Issued on\s*:\s*(\d{2}\s\w+\s\d{4})", text)
    rekening = re.search(r"Nomor Rekening/Account Number\s*:\s*(\d+)", text)
    mata_uang = re.search(r"Mata Uang/Currency\s*:\s*(\w+)", text)
    saldo_awal = re.search(r"Saldo Awal/Initial Balance\s*:\s*([0-9.,+-]+)", text)

    masuk = re.search(
        r"Dana\s+Masuk/Incoming\s+Transactions\s*:\s*([+-]?\s*[\d.,]+)",
        text,
        re.IGNORECASE,
    )
    keluar = re.search(r"Dana Keluar/Outgoing Transactions\s*:\s*([0-9.,+-]+)", text)
    saldo_akhir = re.search(r"Saldo Akhir/Closing Balance\s*:\s*([0-9.,+-]+)", text)

    # Append formatted header lines
    # header_lines.append("E-Statement")
    # header_lines.append("Jl. Maju Jaya, No 1, Jakarta")

    header_lines.append(f"Nama/Name : {name.group(1) if name else '-'}")
    header_lines.append(f"Periode/Period : {periode.group(1)} - {periode.group(2)}" if periode else "-")

    header_lines.append(f"Cabang/Branch : {branch.group(1) if branch else '-'}")
    header_lines.append(f"Dicetak pada/Issued on : {printed.group(1) if printed else '-'}")

    header_lines.append("Tabungan Mandiri")
    header_lines.append(f"Saldo Awal/Initial Balance : {saldo_awal.group(1) if saldo_awal else '-'}")
    header_lines.append(f"Nomor Rekening/Account Number : {rekening.group(1) if rekening else '-'}")
    header_lines.append(f"Mata Uang/Currency : {mata_uang.group(1) if mata_uang else '-'}")

    header_lines.append(f"Dana Masuk/Incoming Transactions : {masuk.group(1) if masuk else '-'}")
    header_lines.append(f"Dana Keluar/Outgoing Transactions : {keluar.group(1) if keluar else '-'}")
    header_lines.append(f"Saldo Akhir/Closing Balance : {saldo_akhir.group(1) if saldo_akhir else '-'}")

    tables = camelot.read_pdf(
        filepath=pdf_path,
        pages="all",
        flavor="stream",
        strip_text="\n",
        edge_tol=300,
        row_tol=10,
    )

    if not tables or tables.n == 0:
        pd.DataFrame({"Error": ["No table data found."]}).to_excel(
            output, index=False, engine="openpyxl"
        )
        output.seek(0)
        return output

    all_dfs = []

    column_keywords = {
        "NO": ["NO", "NO."],
        "TANGGAL": ["TANGGAL", "DATE"],
        "KETERANGAN": ["KETERANGAN", "REMARKS", "URAIAN"],
        "NOMINAL": ["NOMINAL", "AMOUNT"],
        "SALDO": ["SALDO", "BALANCE"],
    }
    keyword_to_standard_col = {
        keyword: standard_col
        for standard_col, keywords in column_keywords.items()
        for keyword in keywords
    }

    for i, table in enumerate(tables[1:-1]):

        df = table.df.copy()
        if df.empty:
            continue

        header_row_candidate_idx = -1
        col_idx_to_standard_name = {}

        for r_idx in range(min(5, df.shape[0])):
            row_values = [
                str(v).upper().replace("\n", " ").strip() for v in df.iloc[r_idx]
            ]
            temp_map = {}
            match_count = 0
            for c_idx, val in enumerate(row_values):
                for keyword, std_col in keyword_to_standard_col.items():
                    if keyword in val:
                        temp_map[c_idx] = std_col
                        match_count += 1
                        break
            if match_count >= 5:
                header_row_candidate_idx = r_idx
                col_idx_to_standard_name = temp_map
                break

        new_cols = [f"Col_{i}" for i in range(df.shape[1])]

        # Ensure unique column names before assignment
        seen = set()
        for idx, name in col_idx_to_standard_name.items():
            original_name = name
            count = 1
            unique_name = name
            while unique_name in seen:
                unique_name = f"{original_name}_{count}"
                count += 1
            seen.add(unique_name)
            new_cols[idx] = unique_name

        df.columns = new_cols

        if header_row_candidate_idx != -1:
            df = df.iloc[header_row_candidate_idx + 1 :].reset_index(drop=True)

        # Remove rows with known non-transactional keywords
        df = df[
            ~df.apply(
                lambda row: row.astype(str)
                .str.contains(
                    r"SALDO|DISCLAIMER|LPS|DITERBITKAN|PERIODE|CATATAN|STATEMENT|DOKUMEN|KEBERATAN|SYARAT|KETENTUAN|E-STATEMENT",
                    case=False,
                )
                .any(),
                axis=1,
            )
        ]

        # Drop rows where all columns are empty or NaN
        df = df.dropna(how="all")
        df = df[~(df == "").all(axis=1)]

        for col in column_keywords.keys():
            if col not in df.columns:
                df[col] = ""

        df.rename(
            columns={
                "NO": "No",
                "TANGGAL": "Tanggal Transaksi",
                "KETERANGAN": "Keterangan Utama",
                "NO_1": "Nominal",
                "SALDO": "Saldo",
            },
            inplace=True,
        )

        df[["Uang Masuk", "Uang Keluar"]] = df.apply(
            classify_mandiri_transaction, axis=1
        )

        final_columns = [
            col
            for col in [
                "No",
                "Tanggal Transaksi",
                "Keterangan Utama",
                "Nominal",
                "Uang Masuk",
                "Uang Keluar",
                "Saldo",
            ]
            if col in df.columns
        ]
        df = df[final_columns]
        
            # Clean Saldo
        if "Saldo" in df.columns:
            df["Saldo"] = (
                df["Saldo"]
                .astype(str)
                .str.replace(".", "")
                .str.replace(",", ".")
                .replace(["<NA>", "nan", "NaN"], "")
                .fillna("")
            )
        
        # Drop optional
        df.drop(
            columns=[
                col for col in ["Nominal"] if col in df.columns
            ],
            inplace=True,
        )

        all_dfs.append(df)

    if not all_dfs:
        pd.DataFrame({"Error": ["No valid data found."]}).to_excel(
            output, index=False, engine="openpyxl"
        )
        output.seek(0)
        return output

    merged_df = pd.concat(all_dfs, ignore_index=True)
    # --- Write to Excel ---

    if export_type == "excel":
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Transaksi"

            # Write header lines
            for i, line in enumerate(header_lines, start=1):
                ws.merge_cells(
                    start_row=i,
                    start_column=1,
                    end_row=i,
                    end_column=merged_df.shape[1],
                )
                cell = ws.cell(row=i, column=1, value=line)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Add table headers and data
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            start_row = len(header_lines) + 2 if header_lines else 1
            for r_idx, row in enumerate(
                dataframe_to_rows(merged_df, index=False, header=True), start=start_row
            ):
                for c_idx, val in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.border = thin_border
                    if r_idx == start_row:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                    else:
                        col_letter = get_column_letter(c_idx)
                        if row[0] == "No" or col_letter == get_column_letter(
                            merged_df.columns.get_loc("No")+1
                        ):
                            cell.alignment = Alignment(
                                horizontal="center", vertical="center"
                            )
                        elif row[
                            1
                        ] == "Tanggal Transaksi" or col_letter == get_column_letter(
                            merged_df.columns.get_loc("Tanggal Transaksi") + 1
                        ):
                            cell.alignment = Alignment(
                                horizontal="left", vertical="center"
                            )
                        elif row[3] == "Uang Masuk" or col_letter == get_column_letter(
                            merged_df.columns.get_loc("Uang Masuk") + 1
                        ):
                            cell.alignment = Alignment(
                                horizontal="right", vertical="top"
                            )
                        elif row[4] == "Uang Keluar" or col_letter == get_column_letter(
                            merged_df.columns.get_loc("Uang Keluar") + 1
                        ):
                            cell.alignment = Alignment(
                                horizontal="right", vertical="top"
                            )
                        elif row[5] == "Saldo" or col_letter == get_column_letter(
                            merged_df.columns.get_loc("Saldo") + 1
                        ):
                            cell.alignment = Alignment(
                                horizontal="right", vertical="top"
                            )
                        else:
                            cell.alignment = Alignment(
                                horizontal="left", vertical="top"
                            )

                        for c_idx, val in enumerate(row, start=1):
                            ws.cell(row=r_idx, column=c_idx, value=val)

            # Auto column width
            no_idx = merged_df.columns.get_loc("No") + 1
            no_idx_letter = get_column_letter(no_idx)

            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)

                # Calculate max width based on actual content length
                max_length = max(
                    len(str(cell.value)) if cell.value else 3 for cell in col
                )

                # Cap the width for "Tanggal Transaksi" to a reasonable size
                if col_letter == no_idx_letter:
                    ws.column_dimensions[col_letter].width = max(
                        15, min(15, max_length + 2)
                    )
                else:
                    ws.column_dimensions[col_letter].width = max_length + 2

            wb.save(output)
            output.seek(0)

        except Exception as e:
            print("Excel export failed:", e)
            output = BytesIO()
            pd.DataFrame({"Error": [str(e)]}).to_excel(
                output, index=False, engine="openpyxl"
            )
            output.seek(0)
    elif export_type == "csv":
        try:
            # Optionally convert specific columns to string to ensure left alignment when opened in Excel
            merged_df["Tanggal Transaksi"] = merged_df["Tanggal Transaksi"].astype(str)
            merged_df["Keterangan Utama"] = merged_df["Keterangan Utama"].astype(str)
            merged_df["Keterangan Tambahan"] = merged_df["Keterangan Tambahan"].astype(
                str
            )

            # Optionally include header_lines as fake "rows" above the CSV data
            csv_string = ""
            if header_lines:
                for line in header_lines:
                    csv_string += (
                        f'"{line}"' + "," * (len(merged_df.columns) - 1) + "\n"
                    )
            csv_string += merged_df.to_csv(index=False)

            output.write(csv_string.encode("utf-8"))
            output.seek(0)

        except Exception as e:
            print("CSV export failed:", e)
            output = BytesIO()
            pd.DataFrame({"Error": [str(e)]}).to_csv(output, index=False)
            output.seek(0)
    else:
        print("Invalid export type")
        output = BytesIO()
        pd.DataFrame({"Error": ["Invalid export type"]}).to_excel(
            output, index=False, engine="openpyxl"
        )
    return output
