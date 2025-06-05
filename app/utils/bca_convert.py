import pdfplumber
from fastapi.responses import StreamingResponse
import camelot
import re
import pandas as pd
from io import BytesIO
from PyPDF2 import PdfReader
from datetime import datetime
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment


def classify_bca_transaction(row):
    try:
        # Clean and convert mutasi to float
        amount = float(str(row["Mutasi"]).replace(",", "").replace(".00", "").strip())
    except:
        return pd.Series(["", ""])  # Invalid amount, return empty

    desc = str(row.get("Keterangan Utama", "")).upper()

    if any(keyword in desc for keyword in ["DB", "DEBIT"]):
        return pd.Series(["", amount])  # Uang Keluar
    elif any(keyword in desc for keyword in ["CR", "CREDIT", "KR OTOMATIS"]):
        return pd.Series([amount, ""])  # Uang Masuk
    elif not desc.strip():
        return pd.Series(["", amount])  # Empty desc, assume Uang Keluar
    else:
        return pd.Series(["", amount])


# Condition when show list all data in one sheet
async def extract_bca_transactions(pdf_path: str, export_type: str) -> BytesIO:
    # --- Extract header and year from PDF ---
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = pdf.pages[0].extract_text()

        lines = text.split("\n")
        header_lines = []

        for line in lines:
            # Remove substring like "HALAMAN : 1 / 2" (case-insensitive)
            cleaned_line = re.sub(
                r"\bHALAMAN\s*:\s*\d+\s*/\s*\d+\b", "", line, flags=re.IGNORECASE
            ).strip()
            header_lines.append(cleaned_line)

            if "PERIODE" in cleaned_line.upper():
                break

        match = re.search(r"PERIODE\s*:\s*(\w+)\s+(\d{4})", text, re.IGNORECASE)
        year_from_periode = match.group(2) if match else str(datetime.today().year)

    except Exception as e:
        print("Header parsing failed:", e)
        header_lines = []
        year_from_periode = str(datetime.today().year)

    # --- Extract tables ---
    tables = camelot.read_pdf(
        filepath=pdf_path,
        pages="all",
        flavor="stream",
        strip_text="\n",
        edge_tol=500,
    )

    if not tables or tables.n == 0:
        print("No tables found in the PDF.")
        output = BytesIO()
        pd.DataFrame({"Error": ["No table data found."]}).to_excel(
            output, index=False, engine="openpyxl"
        )
        output.seek(0)
        return output

    all_dfs = []

    # --- Column mapping ---
    column_keywords = {
        "TANGGAL": ["TANGGAL", "DATE"],
        "KETERANGAN": ["KETERANGAN", "DESCRIPTION", "DETAIL"],
        "CBG": ["CBG", "BRANCH"],
        "MUTASI": ["MUTASI", "DEBIT", "CREDIT", "AMOUNT"],
        "SALDO": ["SALDO", "BALANCE"],
    }

    keyword_to_standard_col = {
        keyword: standard_col
        for standard_col, keywords in column_keywords.items()
        for keyword in keywords
    }

    # --- Process each table ---
    for table in tables:
        df = table.df.copy()
        if df.empty:
            continue

        # --- Identify header row ---
        header_row_candidate_idx = -1
        col_idx_to_standard_name = {}

        for r_idx in range(min(df.shape[0], 5)):
            row_values = [
                str(val).upper().replace("\n", " ").strip() for val in df.iloc[r_idx]
            ]
            temp_col_map = {}
            match_count = 0
            for c_idx, cell_value in enumerate(row_values):
                for keyword, standard_col in keyword_to_standard_col.items():
                    if keyword in cell_value:
                        temp_col_map[c_idx] = standard_col
                        match_count += 1
                        break
            if match_count >= 3:
                header_row_candidate_idx = r_idx
                col_idx_to_standard_name = temp_col_map
                break

        new_df_columns = [f"Col_{i}" for i in range(df.shape[1])]
        for idx, name in col_idx_to_standard_name.items():
            new_df_columns[idx] = name
        df.columns = new_df_columns

        if header_row_candidate_idx != -1:
            df = df.iloc[header_row_candidate_idx + 1 :]

        # Drop unused Col_X
        df = df.loc[
            :, ~df.columns.str.startswith("Col_") | (df.applymap(str).ne("").any())
        ].copy()

        # Fill missing expected columns
        for col in column_keywords:
            if col not in df.columns:
                df[col] = ""

        # Filter bad rows
        if "TANGGAL" in df.columns:
            df = df[
                ~df["TANGGAL"]
                .astype(str)
                .str.contains(r"SALDO AWAL|HALAMAN|Bersambung", na=False)
            ]

        # Rename to final columns
        rename_map = {
            "TANGGAL": "Tanggal Transaksi",
            "Col_1": "Keterangan Utama",
            "KETERANGAN": "Keterangan Tambahan",
            "CBG": "CBG",
            "MUTASI": "Mutasi",
            "Col_5": "Type",
            "SALDO": "Saldo",
        }
        df.rename(columns=rename_map, inplace=True)

        # Format date with year
        def format_date(val):
            try:
                return datetime.strptime(
                    f"{val.strip()}/{year_from_periode}", "%d/%m/%Y"
                ).strftime("%d/%m/%Y")
            except:
                return ""

        if "Tanggal Transaksi" in df.columns:
            df["Tanggal Transaksi"] = (
                df["Tanggal Transaksi"].astype(str).apply(format_date)
            )

        # Add Uang Masuk / Uang Keluar
        df[["Uang Masuk", "Uang Keluar"]] = df.apply(classify_bca_transaction, axis=1)

        # Clean Saldo
        if "Saldo" in df.columns:
            df["Saldo"] = (
                df["Saldo"]
                .astype(str)
                .str.replace(",", "")
                .replace(["<NA>", "nan", "NaN"], "")
                .fillna("")
            )

        # Reorder columns if needed
        if "Saldo" in df.columns:
            saldo_idx = df.columns.get_loc("Saldo")
            umasuk = df.pop("Uang Masuk")
            ukeluar = df.pop("Uang Keluar")
            df.insert(saldo_idx, "Uang Masuk", umasuk)
            df.insert(saldo_idx + 1, "Uang Keluar", ukeluar)

        # Drop optional
        df.drop(
            columns=[
                col for col in ["CBG", "Mutasi", "Type", "Col_6"] if col in df.columns
            ],
            inplace=True,
        )

        all_dfs.append(df)

    # Combine all tables
    if not all_dfs:
        output = BytesIO()
        pd.DataFrame({"Error": ["No valid data found."]}).to_excel(
            output, index=False, engine="openpyxl"
        )
        output.seek(0)
        return output

    merged_df = pd.concat(all_dfs, ignore_index=True)

    # --- Write to Excel ---
    output = BytesIO()
    if export_type == "excel":
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Transaksi"

            # Write header lines
            for i, line in enumerate(header_lines, start=1):
                ws.merge_cells(
                    start_row=i,
                    start_column=1,
                    end_row=i,
                    end_column=merged_df.shape[1],
                )
                cell = ws.cell(row=i, column=1, value=line)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add table headers and data
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            start_row = len(header_lines) + 2 if header_lines else 1
            for r_idx, row in enumerate(
                dataframe_to_rows(merged_df, index=False, header=True), start=start_row
            ):
                for c_idx, val in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.border = thin_border
                    if r_idx == start_row:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                    else:
                        col_letter = get_column_letter(c_idx)
                        if row[
                            0
                        ] == "Tanggal Transaksi" or col_letter == get_column_letter(
                            merged_df.columns.get_loc("Tanggal Transaksi") + 1
                        ):
                            cell.alignment = Alignment(
                                horizontal="left", vertical="center"
                            )
                        elif row[3] == "Uang Masuk" or col_letter == get_column_letter(
                            merged_df.columns.get_loc("Uang Masuk") + 1
                        ):
                            cell.alignment = Alignment(
                                horizontal="right", vertical="top"
                            )
                        elif row[4] == "Uang Keluar" or col_letter == get_column_letter(
                            merged_df.columns.get_loc("Uang Keluar") + 1
                        ):
                            cell.alignment = Alignment(
                                horizontal="right", vertical="top"
                            )
                        else:
                            cell.alignment = Alignment(
                                horizontal="left", vertical="top"
                            )

                        for c_idx, val in enumerate(row, start=1):
                            ws.cell(row=r_idx, column=c_idx, value=val)

            # Auto column width
            tanggal_transaksi_idx = merged_df.columns.get_loc("Tanggal Transaksi") + 1
            tanggal_transaksi_letter = get_column_letter(tanggal_transaksi_idx)

            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)

                # Calculate max width based on actual content length
                max_length = max(
                    len(str(cell.value)) if cell.value else 3 for cell in col
                )

                # Cap the width for "Tanggal Transaksi" to a reasonable size
                if col_letter == tanggal_transaksi_letter:
                    ws.column_dimensions[col_letter].width = max(
                        15, min(15, max_length + 2)
                    )
                else:
                    ws.column_dimensions[col_letter].width = max_length + 2

            wb.save(output)
            output.seek(0)

        except Exception as e:
            print("Excel export failed:", e)
            output = BytesIO()
            pd.DataFrame({"Error": [str(e)]}).to_excel(
                output, index=False, engine="openpyxl"
            )
            output.seek(0)
    elif export_type == "csv":
        try:
            # Optionally convert specific columns to string to ensure left alignment when opened in Excel
            merged_df["Tanggal Transaksi"] = merged_df["Tanggal Transaksi"].astype(str)
            merged_df["Keterangan Utama"] = merged_df["Keterangan Utama"].astype(str)
            merged_df["Keterangan Tambahan"] = merged_df["Keterangan Tambahan"].astype(str)

            # Optionally include header_lines as fake "rows" above the CSV data
            csv_string = ""
            if header_lines:
                for line in header_lines:
                    csv_string += f'"{line}"' + "," * (len(merged_df.columns) - 1) + "\n"
            csv_string += merged_df.to_csv(index=False)

            output.write(csv_string.encode("utf-8"))
            output.seek(0)

        except Exception as e:
            print("CSV export failed:", e)
            output = BytesIO()
            pd.DataFrame({"Error": [str(e)]}).to_csv(output, index=False)
            output.seek(0)
    else:
        print("Invalid export type")
        output = BytesIO()
        pd.DataFrame({"Error": ["Invalid export type"]}).to_excel(
            output, index=False, engine="openpyxl"
        )
    return output


# Condition when sprate with different sheet name
# def extract_bca_transactions(
#     pdf_path: str, bank_type: str, export_type: str
# ) -> BytesIO:
#     column_keywords = {
#         "TANGGAL": ["TANGGAL", "DATE"],
#         "KETERANGAN": ["KETERANGAN", "DESCRIPTION", "DETAIL"],
#         "CBG": ["CBG", "BRANCH"],
#         "MUTASI": ["MUTASI", "DEBIT", "CREDIT", "AMOUNT"],
#         "SALDO": ["SALDO", "BALANCE"],
#     }
#     keyword_to_standard_col = {
#         keyword: standard_col
#         for standard_col, keywords in column_keywords.items()
#         for keyword in keywords
#     }

#     output = BytesIO()
#     wb = Workbook()
#     wb.remove(wb.active)

#     with pdfplumber.open(pdf_path) as pdf:
#         for page_num, page in enumerate(pdf.pages, start=1):
#             text = page.extract_text()
#             lines = text.split("\n") if text else []
#             header_lines = []
#             for line in lines:
#                 header_lines.append(line)
#                 if "PERIODE" in line:
#                     break

#             match = re.search(
#                 r"PERIODE\s*:\s*(\w+)\s+(\d{4})", text or "", re.IGNORECASE
#             )
#             year_from_periode = match.group(2) if match else str(datetime.today().year)

#             tables = camelot.read_pdf(
#                 filepath=pdf_path,
#                 pages=str(page_num),
#                 flavor="stream",
#                 strip_text="\n",
#                 edge_tol=500,
#             )

#             if not tables or tables.n == 0:
#                 continue

#             for table in tables:
#                 df = table.df.copy()
#                 if df.empty:
#                     continue

#                 header_row_idx = -1
#                 col_idx_map = {}

#                 for r_idx in range(min(df.shape[0], 5)):
#                     row_values = [
#                         str(val).upper().replace("\n", " ").strip()
#                         for val in df.iloc[r_idx]
#                     ]
#                     temp_map = {}
#                     match_count = 0
#                     for c_idx, val in enumerate(row_values):
#                         for keyword, standard_col in keyword_to_standard_col.items():
#                             if keyword in val:
#                                 temp_map[c_idx] = standard_col
#                                 match_count += 1
#                                 break
#                     if match_count >= 3:
#                         header_row_idx = r_idx
#                         col_idx_map = temp_map
#                         break

#                 new_columns = [f"Col_{i}" for i in range(df.shape[1])]
#                 for idx, name in col_idx_map.items():
#                     new_columns[idx] = name
#                 df.columns = new_columns

#                 if header_row_idx != -1:
#                     df = df.iloc[header_row_idx + 1 :]

#                 for col in column_keywords:
#                     if col not in df.columns:
#                         df[col] = ""

#                 if "TANGGAL" in df.columns:
#                     df = df[
#                         ~df["TANGGAL"]
#                         .astype(str)
#                         .str.contains(r"SALDO AWAL|HALAMAN|Bersambung", na=False)
#                     ]

#                 df.rename(
#                     columns={
#                         "TANGGAL": "Tanggal Transaksi",
#                         "Col_1": "Keterangan Utama",
#                         "KETERANGAN": "Keterangan Tambahan",
#                         "CBG": "CBG",
#                         "MUTASI": "Mutasi",
#                         "Col_5": "Type",
#                         "SALDO": "Saldo",
#                     },
#                     inplace=True,
#                 )

#                 def format_date(val):
#                     try:
#                         return datetime.strptime(
#                             f"{val.strip()}/{year_from_periode}", "%d/%m/%Y"
#                         ).strftime("%d/%m/%Y")
#                     except:
#                         return ""

#                 if "Tanggal Transaksi" in df.columns:
#                     df["Tanggal Transaksi"] = (
#                         df["Tanggal Transaksi"].astype(str).apply(format_date)
#                     )

#                 df[["Uang Masuk", "Uang Keluar"]] = df.apply(
#                     classify_bca_transaction, axis=1
#                 )

#                 if "Saldo" in df.columns:
#                     df["Saldo"] = (
#                         df["Saldo"]
#                         .astype(str)
#                         .str.replace(",", "")
#                         .replace(["<NA>", "nan", "NaN"], "")
#                         .fillna("")
#                     )

#                 if "Saldo" in df.columns:
#                     idx = df.columns.get_loc("Saldo")
#                     masuk = df.pop("Uang Masuk")
#                     keluar = df.pop("Uang Keluar")
#                     df.insert(idx, "Uang Masuk", masuk)
#                     df.insert(idx + 1, "Uang Keluar", keluar)

#                 df.drop(
#                     columns=[
#                         col
#                         for col in ["CBG", "Mutasi", "Type", "Col_6"]
#                         if col in df.columns
#                     ],
#                     inplace=True,
#                 )

#                 sheet = wb.create_sheet(title=f"Halaman {page_num}")

#                 for i, line in enumerate(header_lines, start=1):
#                     sheet.merge_cells(
#                         start_row=i, start_column=1, end_row=i, end_column=df.shape[1]
#                     )
#                     cell = sheet.cell(row=i, column=1, value=line)
#                     cell.font = Font(bold=True)
#                     cell.alignment = Alignment(horizontal="center")

#                 start_row = len(header_lines) + 2
#                 for r_idx, row in enumerate(
#                     dataframe_to_rows(df, index=False, header=True), start=start_row
#                 ):
#                     for c_idx, val in enumerate(row, start=1):
#                         cell = sheet.cell(row=r_idx, column=c_idx, value=val)
#                         cell.border = Border(
#                             left=Side(style="thin"),
#                             right=Side(style="thin"),
#                             top=Side(style="thin"),
#                             bottom=Side(style="thin"),
#                         )
#                         if r_idx == start_row:
#                             cell.font = Font(bold=True)
#                             cell.alignment = Alignment(horizontal="center")
#                         else:
#                             if df.columns[c_idx - 1] == "Tanggal Transaksi":
#                                 cell.alignment = Alignment(horizontal="center")
#                             elif df.columns[c_idx - 1] in ["Uang Masuk", "Uang Keluar"]:
#                                 cell.alignment = Alignment(horizontal="right")
#                             else:
#                                 cell.alignment = Alignment(horizontal="left")

#                 for col in sheet.columns:
#                     max_length = max(
#                         len(str(cell.value)) if cell.value else 3 for cell in col
#                     )
#                     col_letter = get_column_letter(col[0].column)
#                     if df.columns[col[0].column - 1] == "Tanggal Transaksi":
#                         sheet.column_dimensions[col_letter].width = 15
#                     else:
#                         sheet.column_dimensions[col_letter].width = max(
#                             10, max_length + 2
#                         )

#     wb.save(output)
#     output.seek(0)
#     return output

# original code
# def extract_bca_transactions(pdf_path: str, bank_type: str, export_type: str) -> BytesIO:

#     try:
#         text = pdfplumber.open(pdf_path).pages[0].extract_text()
#         lines = text.split("\n")
#         header_lines = []
#         for line in lines:
#             if "PERIODE" in line:
#                 header_lines.append(line)
#                 break
#             header_lines.append(line)

#         periode_match = re.search(r"PERIODE\s*:\s*(\w+)\s+(\d{4})", text, re.IGNORECASE)
#         year_from_periode = periode_match.group(2) if periode_match else str(datetime.today().year)
#     except:
#         year_from_periode = str(datetime.today().year)
#         header_lines = []

#     tables = camelot.read_pdf(
#         filepath=pdf_path,
#         pages="all",
#         flavor="stream",
#         strip_text="\n",
#         edge_tol=500,
#     )

#     if not tables:
#         print("No tables found in the PDF. Please check the PDF path and structure.")
#         output = BytesIO()
#         pd.DataFrame().to_excel(output, index=False, engine="openpyxl")
#         output.seek(0)
#         return output

#     all_dfs = []

#     column_keywords = {
#         'TANGGAL': ['TANGGAL', 'DATE'],
#         'KETERANGAN': ['KETERANGAN', 'DESCRIPTION', 'DETAIL'],
#         'CBG': ['CBG', 'BRANCH'],
#         'MUTASI': ['MUTASI', 'DEBIT', 'CREDIT', 'AMOUNT'],
#         'SALDO': ['SALDO', 'BALANCE']
#     }

#     keyword_to_standard_col = {
#         keyword: standard_col
#         for standard_col, keywords in column_keywords.items()
#         for keyword in keywords
#     }

#     for i, table in enumerate(tables):
#         df = table.df.copy()
#         if df.shape[0] == 0:
#             continue

#         col_idx_to_standard_name = {}
#         header_row_candidate_idx = -1

#         for r_idx in range(min(df.shape[0], 5)):
#             row_values = [str(val).upper().replace('\n', ' ').strip() for val in df.iloc[r_idx]]
#             found_keywords_count = 0
#             temp_col_map = {}
#             for c_idx, cell_value in enumerate(row_values):
#                 for keyword, standard_col in keyword_to_standard_col.items():
#                     if keyword in cell_value:
#                         temp_col_map[c_idx] = standard_col
#                         found_keywords_count += 1
#                         break
#             if found_keywords_count >= 3:
#                 col_idx_to_standard_name = temp_col_map
#                 header_row_candidate_idx = r_idx
#                 break

#         new_df_columns = [f'Col_{j}' for j in range(df.shape[1])]
#         for c_idx, standard_name in col_idx_to_standard_name.items():
#             if c_idx < len(new_df_columns):
#                 new_df_columns[c_idx] = standard_name

#         df.columns = new_df_columns

#         if header_row_candidate_idx != -1:
#             df = df[header_row_candidate_idx + 1:].copy()

#         df = df.loc[:, ~df.columns.str.startswith('Col_') | (df.apply(lambda x: x.astype(str).str.strip() != '').any())].copy()

#         for col in column_keywords.keys():
#             if col not in df.columns:
#                 df[col] = ''

#         if 'TANGGAL' in df.columns:
#             df = df[~df['TANGGAL'].astype(str).str.contains(r'^(?:SALDO AWAL|HALAMAN|Bersambung)', na=False, regex=True)]

#         df.replace('', pd.NA, inplace=True)
#         df.dropna(subset=list(column_keywords.keys()), how='all', inplace=True)

#         # Custom column remapping
#         rename_map = {
#             'TANGGAL': 'Tanggal Transaksi',
#             'Col_1': 'Keterangan Utama',
#             'KETERANGAN': 'Keterangan Tambahan',
#             'CBG': 'CBG',
#             'MUTASI': 'Mutasi',
#             'Col_5': 'Type',
#             'SALDO': 'Saldo'
#         }
#         df = df.rename(columns=rename_map)

#         df[['Uang Masuk', 'Uang Keluar']] = df.apply(classify_bca_transaction, axis=1)

#         if 'Saldo' in df.columns:
#             saldo_index = df.columns.get_loc('Saldo')
#             # Remove columns temporarily
#             uang_masuk_series = df.pop('Uang Masuk')
#             uang_keluar_series = df.pop('Uang Keluar')
#             # Reinsert at correct position
#             df.insert(saldo_index, 'Uang Masuk', uang_masuk_series)
#             df.insert(saldo_index + 1, 'Uang Keluar', uang_keluar_series)

#         if "Tanggal Transaksi" in df.columns:

#             def add_year_and_format(date_str):
#                 try:
#                     return datetime.strptime(
#                         f"{date_str.strip()}/{year_from_periode}", "%d/%m/%Y"
#                     ).strftime("%d/%m/%Y")
#                 except:
#                     return ""  # If it fails, return as-is

#             df["Tanggal Transaksi"] = (
#                 df["Tanggal Transaksi"].astype(str).apply(add_year_and_format)
#             )

#         if 'Type' in df.columns:
#             df.drop(columns=['Type'], inplace=True)

#         if 'Saldo' in df.columns:
#             df["Saldo"] = (
#                 df["Saldo"]
#                 .astype(str)
#                 .str.replace(",", "", regex=False)
#                 .replace(["<NA>", "nan", "NaN"], "")
#                 .replace({pd.NA: "", None: ""})
#             )

#         if'CBG' in df.columns:
#             df.drop(columns=['CBG'], inplace=True)

#         if 'Col_6' in df.columns:
#             df.drop(columns=['Col_6'], inplace=True)

#             final_columns = [
#                 "Tanggal Transaksi",
#                 "Keterangan Utama",
#                 "Keterangan Tambahan",
#                 "Mutasi",
#                 "Uang Masuk",
#                 "Uang Keluar",
#                 "Type",
#                 "Saldo",
#             ]

#             df = df[[col for col in final_columns if col in df.columns]]

#         all_dfs.append(df)

#     if not all_dfs:
#         output = BytesIO()
#         pd.DataFrame().to_excel(output, index=False, engine="openpyxl")
#         output.seek(0)
#         return output

#     merged_df = pd.concat(all_dfs, ignore_index=True)
#     merged_df.replace('', pd.NA, inplace=True)
#     merged_df.dropna(how='all', inplace=True)

#     output = BytesIO()
#     try:
#         merged_df.to_excel(
#             excel_writer=output,
#             sheet_name="Transaksi",
#             index=False,
#             startrow=9,
#             engine="openpyxl")
#         output.seek(0)
#     except Exception as e:
#         print(f"Error exporting to Excel BytesIO: {e}")
#         output = BytesIO()
#         pd.DataFrame().to_excel(output, index=False, engine="openpyxl")
#         output.seek(0)

#     return output
